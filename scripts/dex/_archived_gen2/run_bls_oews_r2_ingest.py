#!/usr/bin/env python3
"""BLS OEWS — Occupational Employment and Wage Statistics → R2 historical ingest.

Mirrors BLS's OEWS annual survey publications into Cloudflare R2 as
ZSTD-compressed Parquet, year-partitioned + level-partitioned. ~14 years
× 5 aggregation levels = ~70 R2 partitions of aggregate wage / employment
cells (no individual identity).

Source URL pattern (one combined ZIP per survey year):

    https://www.bls.gov/oes/special-requests/oesm{YY}all.zip

The ZIP contains a single XLSX (`oesm{YY}all/all_data_M_{YYYY}.xlsx`)
with ~400K rows × 32 columns, covering all 5 aggregation levels in one
workbook. The R2 ingest splits levels at transform time via the
AREA_TYPE / I_GROUP discriminators:

    AREA_TYPE=1 + I_GROUP=cross-industry  → national    (~800 rows/year)
    AREA_TYPE=1 + I_GROUP!=cross-industry → industry    (varies)
    AREA_TYPE in (2,3)                    → state       (~40K rows/year)
    AREA_TYPE=4                           → msa         (~150-300K rows/year)
    AREA_TYPE in (5,6)                    → non_metro   (~50-70K rows/year)

R2 layout:

    bls-oews/
      year=YYYY/
        level=national/data.parquet
        level=state/data.parquet
        level=msa/data.parquet
        level=non_metro/data.parquet
        level=industry/data.parquet

Audit ledger: ops.bls_oews_r2_ingest_runs (bls_oews_year, level).
Idempotency basis: HEAD Last-Modified per source ZIP. All 5 level rows
for a given year share the same source ZIP — when re-running, the year
is skipped if its source ZIP's Last-Modified matches the most recent
completed run for any of the 5 level rows of the same year.

Pre-2011 OEWS uses a different file structure (per-level ZIPs:
`oesm{YY}st.zip`, `oesm{YY}nat.zip`, `oesm{YY}ma.zip` without combined
'all') and is OUT OF SCOPE for this ingest.

Usage:
  doppler run --project hq-all --config prd -- \\
    uv run python3 scripts/run_bls_oews_r2_ingest.py --all
  doppler run --project hq-all --config prd -- \\
    uv run python3 scripts/run_bls_oews_r2_ingest.py --years 2024
  doppler run --project hq-all --config prd -- \\
    uv run python3 scripts/run_bls_oews_r2_ingest.py --years 2024 \\
      --max-rows 50000 \\
      --r2-prefix-override 'bls-oews/_smoke/year=2024'
  doppler run --project hq-all --config prd -- \\
    uv run python3 scripts/run_bls_oews_r2_ingest.py --years 2022,2023,2024

See directive ~/Desktop/hq/directives/2026-05-08-bls-oews-r2-ingest.md.
"""

from __future__ import annotations

import argparse
import csv
import logging
import os
import re
import shutil
import sys
import time
import zipfile
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import boto3
import duckdb
import httpx
import openpyxl
import psycopg
from psycopg.types.json import Jsonb


R2_BUCKET = "dex-raw-landing-zone"
USER_AGENT = (
    "data-engine-x/bls-oews-r2-ingest (+mailto:tools@substrate.build)"
)
RETRY_STATUSES = {429, 500, 502, 503, 504}
MAX_RETRIES = 5

# Years in scope: oesm{YY}all.zip combined-zip pattern starts at 2011.
# Pre-2011 used per-level zips with a different directive scope.
DEFAULT_YEARS: tuple[int, ...] = tuple(range(2011, 2025))

LEVELS: tuple[str, ...] = (
    "national", "state", "msa", "non_metro", "industry",
)


def _logger() -> logging.Logger:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)-7s %(message)s",
        datefmt="%Y-%m-%dT%H:%M:%S",
    )
    return logging.getLogger("bls-oews-r2-ingest")


log = _logger()


# --------------------------------------------------------------------------- #
# Env / clients
# --------------------------------------------------------------------------- #


def _required_env(name: str) -> str:
    v = os.environ.get(name)
    if not v:
        raise RuntimeError(f"{name} is not set in the environment.")
    return v


def _r2_client() -> "boto3.client":
    return boto3.client(
        "s3",
        endpoint_url=_required_env("R2_ENDPOINT"),
        aws_access_key_id=_required_env("R2_ACCESS_KEY_ID"),
        aws_secret_access_key=_required_env("R2_SECRET_ACCESS_KEY"),
        region_name="auto",
    )


def _database_url() -> str:
    return _required_env("DEX_DB_URL_POOLED")


def source_url_for(year: int) -> str:
    yy = year % 100
    return f"https://www.bls.gov/oes/special-requests/oesm{yy:02d}all.zip"


# --------------------------------------------------------------------------- #
# HTTP layer
# --------------------------------------------------------------------------- #


def head_url(
    client: httpx.Client, url: str,
) -> tuple[int | None, datetime | None, int]:
    last_exc: Exception | None = None
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            r = client.head(url, follow_redirects=True, timeout=30.0)
            if r.status_code in (301, 302, 404):
                return None, None, r.status_code
            if r.status_code in RETRY_STATUSES:
                wait = min(2 ** attempt, 30)
                log.warning("HEAD %s HTTP %s; retry in %ss",
                            url, r.status_code, wait)
                time.sleep(wait)
                continue
            r.raise_for_status()
            cl = int(r.headers.get("content-length", 0)) or None
            lm_raw = r.headers.get("last-modified")
            lm: datetime | None = None
            if lm_raw:
                try:
                    lm = datetime.strptime(
                        lm_raw, "%a, %d %b %Y %H:%M:%S %Z"
                    ).replace(tzinfo=timezone.utc)
                except ValueError:
                    lm = None
            return cl, lm, r.status_code
        except (httpx.RequestError, httpx.HTTPStatusError) as exc:
            last_exc = exc
            wait = min(2 ** attempt, 30)
            log.warning("HEAD %s error (%s); retry in %ss", url, exc, wait)
            time.sleep(wait)
    raise RuntimeError(f"HEAD failed: {last_exc}")


def download_zip(client: httpx.Client, url: str, dest: Path) -> int:
    last_exc: Exception | None = None
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            written = 0
            with client.stream(
                "GET", url, follow_redirects=True, timeout=1800.0,
            ) as r:
                if r.status_code in RETRY_STATUSES:
                    wait = min(2 ** attempt, 30)
                    log.warning("GET %s HTTP %s; retry in %ss",
                                url, r.status_code, wait)
                    time.sleep(wait)
                    continue
                r.raise_for_status()
                with dest.open("wb") as f:
                    last_log = time.monotonic()
                    for chunk in r.iter_bytes(chunk_size=1 << 20):
                        f.write(chunk)
                        written += len(chunk)
                        now = time.monotonic()
                        if now - last_log >= 15.0:
                            log.info(
                                "  download progress: %.1f MB written",
                                written / (1 << 20),
                            )
                            last_log = now
            return written
        except (httpx.RequestError, httpx.HTTPStatusError) as exc:
            last_exc = exc
            wait = min(2 ** attempt, 30)
            log.warning("GET %s error (%s); retry in %ss", url, exc, wait)
            time.sleep(wait)
    raise RuntimeError(f"download failed: {last_exc}")


# --------------------------------------------------------------------------- #
# ZIP unpacking — locate the single XLSX inside oesm{YY}all/<file>.xlsx
# --------------------------------------------------------------------------- #


def extract_xlsx(zip_path: Path, workdir: Path) -> Path:
    with zipfile.ZipFile(zip_path) as zf:
        names = zf.namelist()
        xlsx_names = [n for n in names if n.lower().endswith(".xlsx")]
        if not xlsx_names:
            raise RuntimeError(f"no .xlsx in {zip_path}: contents={names}")
        if len(xlsx_names) > 1:
            log.warning("multiple xlsx in zip; using first: %s", xlsx_names)
        xlsx_name = xlsx_names[0]
        zf.extract(xlsx_name, workdir)
        xlsx_path = workdir / xlsx_name
    log.info(
        "  extracted %s (%.1f MB)",
        xlsx_path.name, xlsx_path.stat().st_size / (1 << 20),
    )
    return xlsx_path


def xlsx_to_csv(xlsx_path: Path, csv_path: Path) -> tuple[list[str], int]:
    """Stream the OEWS workbook's primary 'All May …' sheet to CSV.

    DuckDB's spatial extension XLSX reader segfaults on some BLS OEWS
    workbooks (observed on 2015+ files). The fallback per directive
    lessons #3 is to use openpyxl as the XLSX reader. We stream rows
    in read-only mode to keep memory bounded and emit a single CSV that
    DuckDB then handles via `read_csv` (which is rock solid).

    Returns: (header_list, data_row_count).
    """
    wb = openpyxl.load_workbook(
        xlsx_path, read_only=True, data_only=True,
    )
    # OEWS workbooks have a primary 'All May YYYY data' sheet; some
    # publications also bundle a 'Field Descriptions' sheet. The
    # primary data sheet is conventionally the active one. Defensive:
    # prefer 'All ...' titled sheets if multi-sheet.
    ws = None
    if len(wb.sheetnames) > 1:
        for name in wb.sheetnames:
            if name.lower().startswith("all"):
                ws = wb[name]
                break
    if ws is None:
        ws = wb.active

    log.info("  openpyxl: reading sheet %r", ws.title)

    rows_iter = ws.iter_rows(values_only=True)
    headers_raw = next(rows_iter, None)
    if not headers_raw:
        wb.close()
        raise RuntimeError(f"empty XLSX: {xlsx_path}")
    # Drop trailing None columns (openpyxl pads ws.max_column when the
    # writer reserved blanks).
    headers: list[str] = []
    last_real = -1
    for i, h in enumerate(headers_raw):
        s = "" if h is None else str(h).strip()
        headers.append(s)
        if s:
            last_real = i
    headers = headers[: last_real + 1]
    n_cols = len(headers)

    csv_path.parent.mkdir(parents=True, exist_ok=True)
    n = 0
    with csv_path.open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(headers)
        last_log = time.monotonic()
        for row in rows_iter:
            if row is None:
                continue
            # Trim to header width + skip wholly blank rows.
            cells = list(row[:n_cols])
            if all(c is None or (isinstance(c, str) and not c.strip())
                   for c in cells):
                continue
            # Coerce floats that are integers to int strings (preserves
            # AREA, NAICS leading digits without trailing '.0').
            out: list[str] = []
            for c in cells:
                if c is None:
                    out.append("")
                elif isinstance(c, float) and c.is_integer():
                    out.append(str(int(c)))
                else:
                    out.append(str(c))
            w.writerow(out)
            n += 1
            now = time.monotonic()
            if now - last_log >= 15.0:
                log.info("    csv progress: %d rows", n)
                last_log = now

    wb.close()
    log.info(
        "  csv written: %s rows → %s (%.1f MB)",
        f"{n:,}", csv_path.name, csv_path.stat().st_size / (1 << 20),
    )
    return headers, n


# --------------------------------------------------------------------------- #
# DuckDB level-split + Parquet build
# --------------------------------------------------------------------------- #


# Canonical OEWS column → tuple of historical aliases (lowercase + normalized).
# BLS OEWS column names drift across years:
#   - Case: 2011/2020+ UPPER, 2012-2014/2019 lower (DuckDB st_read preserves)
#   - Spaces: 2014 has "occ code" (with space), normalized to "occ_code"
#   - Renames: 2011 LOC_Q → LOC_QUOTIENT (2012+); 2011-2012 PCT_TOT →
#              PCT_TOTAL (2013+); 2011-2018 GROUP → O_GROUP (2019+);
#              2019 JOBS_1000_ORIG → JOBS_1000
#   - Additions: PRIM_STATE added 2020+; I_GROUP added 2019+; PCT_RPT
#                added 2022+
#
# Resolution: at read time we lowercase + replace whitespace with '_' on
# every actual column name from the XLSX, then walk this tuple per
# canonical column to find the first matching alias. Columns absent
# from a particular year's XLSX project as `NULL AS canonical_name`.
_OEWS_COLUMN_ALIASES: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("area",          ("area",)),
    ("area_title",    ("area_title",)),
    ("area_type",     ("area_type",)),
    ("prim_state",    ("prim_state",)),
    ("naics",         ("naics",)),
    ("naics_title",   ("naics_title",)),
    ("i_group",       ("i_group",)),
    ("own_code",      ("own_code",)),
    ("occ_code",      ("occ_code",)),
    ("occ_title",     ("occ_title",)),
    ("o_group",       ("o_group", "group")),
    ("tot_emp",       ("tot_emp",)),
    ("emp_prse",      ("emp_prse",)),
    ("jobs_1000",     ("jobs_1000", "jobs_1000_orig")),
    ("loc_quotient",  ("loc_quotient", "loc_q")),
    ("pct_total",     ("pct_total", "pct_tot")),
    ("pct_rpt",       ("pct_rpt",)),
    ("h_mean",        ("h_mean",)),
    ("a_mean",        ("a_mean",)),
    ("mean_prse",     ("mean_prse",)),
    ("h_pct10",       ("h_pct10",)),
    ("h_pct25",       ("h_pct25",)),
    ("h_median",      ("h_median",)),
    ("h_pct75",       ("h_pct75",)),
    ("h_pct90",       ("h_pct90",)),
    ("a_pct10",       ("a_pct10",)),
    ("a_pct25",       ("a_pct25",)),
    ("a_median",      ("a_median",)),
    ("a_pct75",       ("a_pct75",)),
    ("a_pct90",       ("a_pct90",)),
    ("annual",        ("annual",)),
    ("hourly",        ("hourly",)),
)
_OEWS_RAW_COLS: tuple[str, ...] = tuple(c for c, _ in _OEWS_COLUMN_ALIASES)

# Numeric cells that may carry BLS suppression sentinels (`'*'`, `'**'`, `'#'`).
# We preserve the raw VARCHAR + add a typed DOUBLE column with TRY_CAST.
_NUMERIC_COLS: tuple[str, ...] = (
    "TOT_EMP", "EMP_PRSE", "JOBS_1000", "LOC_QUOTIENT", "PCT_TOTAL", "PCT_RPT",
    "H_MEAN", "A_MEAN", "MEAN_PRSE",
    "H_PCT10", "H_PCT25", "H_MEDIAN", "H_PCT75", "H_PCT90",
    "A_PCT10", "A_PCT25", "A_MEDIAN", "A_PCT75", "A_PCT90",
)


# DuckDB macros mirroring scripts/_lib/bls_oews_normalize.py. Pinned by
# tests/scripts/test_bls_oews_normalize_sql_parity.py — when a rule
# changes, update BOTH the macro and the Python ref impl.
_NORMALIZE_MACROS_SQL = r"""
-- 7-char canonical SOC code ('NN-NNNN').
CREATE MACRO bls_oews_normalize_soc(raw) AS (
  CASE
    WHEN raw IS NULL OR trim(CAST(raw AS VARCHAR)) = '' THEN NULL
    WHEN regexp_matches(trim(CAST(raw AS VARCHAR)), '^[0-9]{2}-?[0-9]{4}$')
      THEN substring(regexp_replace(trim(CAST(raw AS VARCHAR)), '-', '', 'g'), 1, 2)
           || '-'
           || substring(regexp_replace(trim(CAST(raw AS VARCHAR)), '-', '', 'g'), 3, 4)
    ELSE NULL
  END
);

-- Digits-only canonical area code.
CREATE MACRO bls_oews_normalize_area_code(raw) AS (
  CASE
    WHEN raw IS NULL OR trim(CAST(raw AS VARCHAR)) = '' THEN NULL
    WHEN length(regexp_replace(CAST(raw AS VARCHAR), '\D', '', 'g')) = 0 THEN NULL
    ELSE regexp_replace(CAST(raw AS VARCHAR), '\D', '', 'g')
  END
);

-- Uppercased + whitespace-collapsed area title.
CREATE MACRO bls_oews_normalize_area_title(raw) AS (
  CASE
    WHEN raw IS NULL OR trim(CAST(raw AS VARCHAR)) = '' THEN NULL
    ELSE upper(trim(regexp_replace(CAST(raw AS VARCHAR), '\s+', ' ', 'g')))
  END
);

-- Digits-only canonical NAICS.
CREATE MACRO bls_oews_normalize_naics(raw) AS (
  CASE
    WHEN raw IS NULL OR trim(CAST(raw AS VARCHAR)) = '' THEN NULL
    WHEN length(regexp_replace(CAST(raw AS VARCHAR), '\D', '', 'g')) = 0 THEN NULL
    ELSE regexp_replace(CAST(raw AS VARCHAR), '\D', '', 'g')
  END
);
"""


def _level_filter_sql(level: str) -> str:
    """Cross-year-stable level filter. Uses NAICS = '000000' to discriminate
    national-cross-industry from national-by-industry, since I_GROUP only
    exists 2019+. State / MSA / non-metro keys off AREA_TYPE alone.

    AREA_TYPE values (BLS, all years):
      1 = U.S. national     2 = State    3 = Territory (PR/VI/GU)
      4 = MSA               5/6 = Non-metropolitan area
    """
    if level == "national":
        return (
            "TRY_CAST(NULLIF(trim(area_type), '') AS INT) = 1 "
            "AND coalesce(trim(naics), '') = '000000'"
        )
    if level == "industry":
        return (
            "TRY_CAST(NULLIF(trim(area_type), '') AS INT) = 1 "
            "AND coalesce(trim(naics), '') <> '000000'"
        )
    if level == "state":
        return "TRY_CAST(NULLIF(trim(area_type), '') AS INT) IN (2, 3)"
    if level == "msa":
        return "TRY_CAST(NULLIF(trim(area_type), '') AS INT) = 4"
    if level == "non_metro":
        return "TRY_CAST(NULLIF(trim(area_type), '') AS INT) IN (5, 6)"
    raise ValueError(f"unknown level {level!r}")


def _normalize_xlsx_col_key(name: str) -> str:
    """Lowercase + collapse whitespace to '_' for case/space-insensitive
    matching. Mirrors how 2014's "occ code" / "occ title" map back to
    canonical "occ_code" / "occ_title"."""
    return re.sub(r"\s+", "_", name.strip().lower())


def _resolve_actual_columns(
    actual_cols: list[str],
) -> dict[str, str | None]:
    """Map canonical OEWS column → actual column name in this XLSX, or
    None if absent. Walks _OEWS_COLUMN_ALIASES per canonical column."""
    keyed = {_normalize_xlsx_col_key(c): c for c in actual_cols}
    out: dict[str, str | None] = {}
    for canonical, aliases in _OEWS_COLUMN_ALIASES:
        out[canonical] = None
        for alias in aliases:
            actual = keyed.get(alias)
            if actual is not None:
                out[canonical] = actual
                break
    return out


def _project_raw_cols_sql(actual_cols: list[str]) -> str:
    """SELECT clause that projects every canonical OEWS column. Maps
    each canonical name to its actual column name in the XLSX (handling
    case + whitespace + historical renames). Missing columns yield NULL
    so all years share the same Parquet shape."""
    resolved = _resolve_actual_columns(actual_cols)
    parts: list[str] = []
    for canonical in _OEWS_RAW_COLS:
        actual = resolved[canonical]
        if actual is None:
            parts.append(f"CAST(NULL AS VARCHAR) AS {canonical}")
        else:
            parts.append(f'CAST("{actual}" AS VARCHAR) AS {canonical}')
    return ",\n  ".join(parts)


def transform_level_to_parquet(
    *,
    csv_path: Path,
    csv_headers: list[str],
    parquet_path: Path,
    bls_oews_year: int,
    level: str,
    soc_revision: str,
    max_rows: int | None,
) -> tuple[int, int]:
    """Read source CSV → filter to one level → write ZSTD Parquet.
    Returns (source_rows_in_level, parquet_rows)."""
    con = duckdb.connect(":memory:")
    con.execute("PRAGMA threads=4;")
    con.execute("PRAGMA memory_limit='6GB';")

    raw_proj = _project_raw_cols_sql(csv_headers)

    # Stage: read the CSV emitted from openpyxl streaming. all_varchar
    # preserves BLS suppression sentinels ('*', '**', '#') in the raw
    # columns; numeric companions are TRY_CAST below.
    con.execute(f"""
        CREATE VIEW raw_xlsx AS
        SELECT
          {raw_proj}
        FROM read_csv(
          '{csv_path}',
          delim=',', header=TRUE,
          all_varchar=TRUE,
          ignore_errors=TRUE, null_padding=TRUE
        );
    """)

    # rows_in_level = how many source rows match this level's filter.
    filter_sql = _level_filter_sql(level)
    rows_in_row = con.execute(
        f"SELECT count(*) FROM raw_xlsx WHERE {filter_sql};"
    ).fetchone()
    rows_in = int(rows_in_row[0]) if rows_in_row else 0

    limit_clause = f"LIMIT {max_rows}" if max_rows is not None else ""

    # Build the typed projection. Casts:
    #   raw VARCHAR (preserved for sentinels) → kept verbatim
    #   {raw}_typed DOUBLE → TRY_CAST sentinels yield NULL
    typed_casts: list[str] = []
    for col in _NUMERIC_COLS:
        lower = col.lower()
        typed_casts.append(
            f"TRY_CAST(NULLIF(trim({lower}), '') AS DOUBLE) AS {lower}_typed"
        )

    parquet_path.parent.mkdir(parents=True, exist_ok=True)

    con.execute(_NORMALIZE_MACROS_SQL)

    typed_proj = ",\n  ".join(typed_casts)

    select_sql = f"""
        SELECT
          -- All canonical OEWS columns preserved as VARCHAR.
          area, area_title, area_type, prim_state,
          naics, naics_title, i_group, own_code,
          occ_code, occ_title, o_group,
          tot_emp, emp_prse, jobs_1000, loc_quotient,
          pct_total, pct_rpt,
          h_mean, a_mean, mean_prse,
          h_pct10, h_pct25, h_median, h_pct75, h_pct90,
          a_pct10, a_pct25, a_median, a_pct75, a_pct90,
          annual, hourly,
          -- Numeric-typed companions (TRY_CAST yields NULL on '*'/'**'/'#' sentinels).
          {typed_proj},
          -- Identity-spine + cross-year columns.
          bls_oews_normalize_soc(occ_code)            AS occ_code_normalized,
          bls_oews_normalize_area_title(occ_title)    AS occ_title_normalized,
          bls_oews_normalize_area_code(area)          AS area_code_normalized,
          bls_oews_normalize_area_title(area_title)   AS area_title_normalized,
          bls_oews_normalize_naics(naics)             AS naics_code_normalized,
          CAST('{level}' AS VARCHAR)                  AS area_kind,
          CAST({bls_oews_year} AS SMALLINT)           AS bls_oews_year,
          CAST('{soc_revision}' AS VARCHAR)           AS soc_revision
        FROM raw_xlsx
        WHERE {filter_sql}
        {limit_clause}
    """

    con.execute(f"""
        COPY ({select_sql}) TO '{parquet_path}'
        (FORMAT PARQUET, COMPRESSION ZSTD, ROW_GROUP_SIZE 100000);
    """)

    rows_pq_row = con.execute(
        f"SELECT count(*) FROM read_parquet('{parquet_path}');"
    ).fetchone()
    rows_pq = int(rows_pq_row[0]) if rows_pq_row else 0

    con.close()
    return rows_in, rows_pq


def parquet_null_rates(parquet_path: Path) -> dict[str, float]:
    con = duckdb.connect(":memory:")
    row = con.execute(f"""
        SELECT
          count(*) AS total,
          count(*) FILTER (WHERE occ_code_normalized IS NULL) AS occ_null,
          count(*) FILTER (WHERE area_code_normalized IS NULL) AS area_null,
          count(*) FILTER (WHERE a_mean IS NULL OR trim(a_mean) = '') AS amean_null
        FROM read_parquet('{parquet_path}');
    """).fetchone()
    con.close()
    total = int(row[0]) if row else 0
    if total == 0:
        return {"occ_code_null_pct": 0.0, "area_code_null_pct": 0.0,
                "a_mean_null_pct": 0.0}
    return {
        "occ_code_null_pct": round(100.0 * int(row[1]) / total, 4),
        "area_code_null_pct": round(100.0 * int(row[2]) / total, 4),
        "a_mean_null_pct": round(100.0 * int(row[3]) / total, 4),
    }


def parquet_column_count(parquet_path: Path) -> int:
    con = duckdb.connect(":memory:")
    rows = con.execute(
        f"DESCRIBE SELECT * FROM read_parquet('{parquet_path}');"
    ).fetchall()
    con.close()
    return len(rows)


# --------------------------------------------------------------------------- #
# R2 + audit helpers
# --------------------------------------------------------------------------- #


def upload_to_r2(parquet_path: Path, *, bucket: str, key: str) -> int:
    s3 = _r2_client()
    file_bytes = parquet_path.stat().st_size
    s3.upload_file(
        str(parquet_path), bucket, key,
        ExtraArgs={"ContentType": "application/x-parquet"},
    )
    return file_bytes


def get_prior_source_last_modified_for_year(
    conn: psycopg.Connection, bls_oews_year: int,
) -> datetime | None:
    """Idempotency key shared across all 5 levels for a given year:
    a year's source ZIP serves all levels."""
    with conn.cursor() as cur:
        cur.execute("""
            SELECT source_last_modified
              FROM ops.bls_oews_r2_ingest_runs
             WHERE bls_oews_year = %s AND status = 'completed'
             ORDER BY started_at DESC LIMIT 1
        """, (bls_oews_year,))
        row = cur.fetchone()
    return row[0] if row else None


def insert_run_row(
    conn: psycopg.Connection,
    *,
    bls_oews_year: int, level: str,
    source_url: str | None,
    source_last_modified: datetime | None,
    prior_source_last_modified: datetime | None,
) -> str:
    sql = """
    INSERT INTO ops.bls_oews_r2_ingest_runs (
        bls_oews_year, level, status,
        source_url, source_last_modified, prior_source_last_modified
    ) VALUES (%s, %s, 'running', %s, %s, %s)
    RETURNING id;
    """
    with conn.cursor() as cur:
        cur.execute(sql, (
            bls_oews_year, level, source_url,
            source_last_modified, prior_source_last_modified,
        ))
        row_id = cur.fetchone()[0]
    conn.commit()
    return str(row_id)


def write_no_change_run(
    conn: psycopg.Connection,
    *,
    bls_oews_year: int, level: str,
    source_url: str | None,
    source_last_modified: datetime | None,
    prior_source_last_modified: datetime | None,
) -> None:
    started = datetime.now(timezone.utc)
    with conn.cursor() as cur:
        cur.execute("""
            INSERT INTO ops.bls_oews_r2_ingest_runs (
                bls_oews_year, level, status,
                source_url, source_last_modified, prior_source_last_modified,
                started_at, finished_at, duration_seconds, notes
            ) VALUES (%s, %s, 'no_change', %s, %s, %s, %s, %s, 0, %s);
        """, (
            bls_oews_year, level, source_url,
            source_last_modified, prior_source_last_modified,
            started, started,
            Jsonb({"reason": "source_last_modified unchanged"}),
        ))
    conn.commit()


def finalize_run_row(
    conn: psycopg.Connection, run_id: str,
    *,
    status: str,
    started_wall: float,
    source_bytes: int | None,
    source_rows: int | None,
    parquet_rows: int | None,
    parquet_bytes: int | None,
    parquet_columns: int | None,
    r2_bucket: str | None,
    r2_prefix: str | None,
    r2_object_key: str | None,
    r2_total_bytes: int | None,
    null_rates: dict[str, float] | None,
    error_message: str | None,
    notes: dict[str, Any] | None,
) -> None:
    duration = round(time.monotonic() - started_wall, 3)
    with conn.cursor() as cur:
        cur.execute("""
            UPDATE ops.bls_oews_r2_ingest_runs
               SET status = %s,
                   source_bytes_downloaded = %s,
                   source_row_count = %s,
                   parquet_row_count = %s,
                   parquet_bytes_written = %s,
                   parquet_column_count = %s,
                   r2_bucket = %s,
                   r2_prefix = %s,
                   r2_object_key = %s,
                   r2_total_bytes = %s,
                   occ_code_null_pct = %s,
                   area_code_null_pct = %s,
                   a_mean_null_pct = %s,
                   finished_at = now(),
                   duration_seconds = %s,
                   error_message = %s,
                   notes = %s
             WHERE id = %s;
        """, (
            status, source_bytes, source_rows,
            parquet_rows, parquet_bytes, parquet_columns,
            r2_bucket, r2_prefix, r2_object_key, r2_total_bytes,
            (null_rates or {}).get("occ_code_null_pct"),
            (null_rates or {}).get("area_code_null_pct"),
            (null_rates or {}).get("a_mean_null_pct"),
            duration, error_message,
            Jsonb(notes) if notes else None, run_id,
        ))
    conn.commit()


# --------------------------------------------------------------------------- #
# Per-year orchestration (5 levels share one downloaded ZIP)
# --------------------------------------------------------------------------- #


@dataclass
class YearContext:
    bls_oews_year: int
    source_url: str
    source_last_modified: datetime | None
    source_bytes: int | None  # bytes-on-wire (from HEAD content-length)


def run_one_year(
    bls_oews_year: int,
    *,
    client: httpx.Client,
    workdir: Path,
    levels: tuple[str, ...],
    max_rows: int | None,
    r2_prefix_override: str | None,
    skip_if_unchanged: bool,
    dry_run: bool,
) -> int:
    log.info("=" * 70)
    log.info("=== INGEST: year=%d levels=%s ===",
             bls_oews_year, ",".join(levels))
    log.info("=" * 70)

    url = source_url_for(bls_oews_year)
    cl, lm, status = head_url(client, url)
    log.info("[%d] HEAD %s → %s  size=%s  last_modified=%s",
             bls_oews_year, url, status, cl, lm)
    if status in (301, 302, 404):
        log.warning("[%d] source missing (HTTP %s) — skipping", bls_oews_year, status)
        # Stamp 'failed' rows for each level so the audit trail reflects "we tried".
        with psycopg.connect(_database_url()) as conn:
            for level in levels:
                run_id = insert_run_row(
                    conn,
                    bls_oews_year=bls_oews_year, level=level,
                    source_url=url, source_last_modified=None,
                    prior_source_last_modified=None,
                )
                finalize_run_row(
                    conn, run_id, status="failed",
                    started_wall=time.monotonic(),
                    source_bytes=None, source_rows=None,
                    parquet_rows=None, parquet_bytes=None,
                    parquet_columns=None,
                    r2_bucket=None, r2_prefix=None, r2_object_key=None,
                    r2_total_bytes=None,
                    null_rates=None,
                    error_message=f"source unavailable HTTP {status}",
                    notes={"http_status": status},
                )
        return 1

    if dry_run:
        log.info("[%d] DRY RUN — exiting after HEAD probe", bls_oews_year)
        return 0

    # Skip-if-unchanged check: a year's source ZIP serves all 5 levels.
    if skip_if_unchanged:
        with psycopg.connect(_database_url()) as conn:
            prior = get_prior_source_last_modified_for_year(conn, bls_oews_year)
            log.info("[%d] prior source_last_modified: %s",
                     bls_oews_year, prior)
            if prior is not None and lm is not None and lm <= prior:
                log.info("[%d] source unchanged — recording no_change for all levels",
                         bls_oews_year)
                for level in levels:
                    write_no_change_run(
                        conn,
                        bls_oews_year=bls_oews_year, level=level,
                        source_url=url, source_last_modified=lm,
                        prior_source_last_modified=prior,
                    )
                return 0

    # Download ZIP once; reuse for all 5 levels.
    zip_path = workdir / f"oesm{bls_oews_year % 100:02d}all.zip"
    log.info("[%d] downloading ZIP → %s", bls_oews_year, zip_path)
    src_bytes = download_zip(client, url, zip_path)
    log.info("[%d] downloaded %.1f MB", bls_oews_year, src_bytes / (1 << 20))

    xlsx_path = extract_xlsx(zip_path, workdir)

    # Convert XLSX → CSV via openpyxl streaming. DuckDB's spatial
    # extension XLSX reader segfaults on some BLS publications (observed
    # 2015+). The CSV-shaped intermediate is rock-solid + lets us share
    # the parsed rows across all 5 level transforms cheaply.
    csv_path = xlsx_path.with_suffix(".csv")
    csv_headers, csv_rows = xlsx_to_csv(xlsx_path, csv_path)
    log.info(
        "[%d] CSV ready: %s data rows, %d columns",
        bls_oews_year, f"{csv_rows:,}", len(csv_headers),
    )

    soc_revision = "SOC2010" if bls_oews_year <= 2017 else "SOC2018"

    # Per-level run loop.
    overall_rc = 0
    for level in levels:
        rc = run_one_level(
            bls_oews_year=bls_oews_year, level=level,
            csv_path=csv_path, csv_headers=csv_headers,
            source_url=url, source_last_modified=lm, source_bytes=src_bytes,
            workdir=workdir,
            max_rows=max_rows,
            r2_prefix_override=r2_prefix_override,
            soc_revision=soc_revision,
        )
        if rc != 0:
            overall_rc = rc
            log.error("[%d/%s] level failed; continuing", bls_oews_year, level)

    # Cleanup: delete extracted XLSX + CSV + zip (~400-500 MB / year).
    try:
        csv_path.unlink(missing_ok=True)
        xlsx_path.unlink(missing_ok=True)
        try:
            xlsx_path.parent.rmdir()
        except OSError:
            pass
        zip_path.unlink(missing_ok=True)
    except OSError:
        pass

    return overall_rc


def run_one_level(
    *,
    bls_oews_year: int, level: str,
    csv_path: Path, csv_headers: list[str],
    source_url: str, source_last_modified: datetime | None,
    source_bytes: int,
    workdir: Path,
    max_rows: int | None,
    r2_prefix_override: str | None,
    soc_revision: str,
) -> int:
    started_wall = time.monotonic()
    log.info("--- [%d/%s] start ---", bls_oews_year, level)

    with psycopg.connect(_database_url()) as conn:
        prior = get_prior_source_last_modified_for_year(conn, bls_oews_year)

        run_id = insert_run_row(
            conn,
            bls_oews_year=bls_oews_year, level=level,
            source_url=source_url,
            source_last_modified=source_last_modified,
            prior_source_last_modified=prior,
        )

        try:
            parquet_path = workdir / f"bls_oews_{bls_oews_year}_{level}.parquet"
            rows_in, rows_pq = transform_level_to_parquet(
                csv_path=csv_path, csv_headers=csv_headers,
                parquet_path=parquet_path,
                bls_oews_year=bls_oews_year, level=level,
                soc_revision=soc_revision, max_rows=max_rows,
            )
            null_rates = parquet_null_rates(parquet_path)
            parquet_columns = parquet_column_count(parquet_path)
            log.info(
                "[%d/%s] transform: source_rows=%s parquet_rows=%s "
                "occ_null=%.2f%% area_null=%.2f%% amean_null=%.2f%%",
                bls_oews_year, level, f"{rows_in:,}", f"{rows_pq:,}",
                null_rates["occ_code_null_pct"],
                null_rates["area_code_null_pct"],
                null_rates["a_mean_null_pct"],
            )

            target_prefix = (
                r2_prefix_override
                or f"bls-oews/year={bls_oews_year}/level={level}"
            )
            target_key = target_prefix.rstrip("/") + "/data.parquet"
            uploaded = upload_to_r2(
                parquet_path, bucket=R2_BUCKET, key=target_key,
            )
            log.info(
                "[%d/%s] uploaded → s3://%s/%s (%.2f MB)",
                bls_oews_year, level, R2_BUCKET, target_key,
                uploaded / (1 << 20),
            )

            finalize_run_row(
                conn, run_id, status="completed",
                started_wall=started_wall,
                source_bytes=source_bytes,
                source_rows=rows_in,
                parquet_rows=rows_pq,
                parquet_bytes=uploaded,
                parquet_columns=parquet_columns,
                r2_bucket=R2_BUCKET,
                r2_prefix=target_prefix.rstrip("/") + "/",
                r2_object_key=target_key,
                r2_total_bytes=uploaded,
                null_rates=null_rates,
                error_message=None,
                notes={
                    "max_rows": max_rows,
                    "r2_prefix_override": r2_prefix_override,
                    "soc_revision": soc_revision,
                    "csv_source_rows": len(csv_headers),
                },
            )
            log.info(
                "[%d/%s] DONE wall=%.1fs",
                bls_oews_year, level,
                time.monotonic() - started_wall,
            )

            try:
                parquet_path.unlink(missing_ok=True)
            except OSError:
                pass
            return 0

        except Exception as exc:
            log.exception("[%d/%s] ingest failed", bls_oews_year, level)
            try:
                finalize_run_row(
                    conn, run_id, status="failed",
                    started_wall=started_wall,
                    source_bytes=None, source_rows=None,
                    parquet_rows=None, parquet_bytes=None,
                    parquet_columns=None,
                    r2_bucket=None, r2_prefix=None, r2_object_key=None,
                    r2_total_bytes=None,
                    null_rates=None,
                    error_message=str(exc)[:1000], notes=None,
                )
            except Exception:
                log.exception("[%d/%s] failed to finalize audit row on error",
                              bls_oews_year, level)
            return 1


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #


def parse_years(spec: str) -> list[int]:
    """Parse a CLI year spec into an ordered list of years.

    Accepts: '2024' | '2022,2023,2024' | '2011-2024' | mixed.
    """
    out: list[int] = []
    for tok in spec.split(","):
        tok = tok.strip()
        if not tok:
            continue
        if "-" in tok:
            a, b = tok.split("-", 1)
            for y in range(int(a), int(b) + 1):
                out.append(y)
        else:
            out.append(int(tok))
    return sorted(set(out))


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument(
        "--years", default=None,
        help="Year spec: '2024' | '2022,2023,2024' | '2011-2024'.",
    )
    p.add_argument("--all", action="store_true",
                   help=f"Run all in-scope years: {DEFAULT_YEARS[0]}-{DEFAULT_YEARS[-1]}.")
    p.add_argument(
        "--levels", default=None,
        help="Comma-separated levels to run. Default: all 5. "
             f"Choices: {','.join(LEVELS)}.",
    )
    p.add_argument("--workdir", default=None,
                   help="Staging directory. Default /tmp/bls_oews_r2_ingest.")
    p.add_argument("--r2-prefix-override", default=None,
                   help="Replace canonical bls-oews/year=*/level=* prefix "
                        "(smoke testing).")
    p.add_argument("--max-rows", type=int, default=None,
                   help="Smoke testing: cap rows per (year,level) Parquet.")
    p.add_argument("--skip-if-unchanged", action="store_true",
                   help="Short-circuit if upstream Last-Modified matches "
                        "prior completed run for the same year.")
    p.add_argument("--dry-run", action="store_true",
                   help="HEAD/probe only; no DB or R2 writes.")
    return p.parse_args()


def main() -> int:
    args = parse_args()

    if args.years:
        years = parse_years(args.years)
    elif args.all:
        years = list(DEFAULT_YEARS)
    else:
        log.error("must pass --years or --all")
        return 2

    if args.levels:
        wanted = tuple(s.strip() for s in args.levels.split(",") if s.strip())
        unknown = set(wanted) - set(LEVELS)
        if unknown:
            log.error("unknown level(s): %s; valid: %s",
                      sorted(unknown), LEVELS)
            return 2
        levels = wanted
    else:
        levels = LEVELS

    workdir = Path(args.workdir or "/tmp/bls_oews_r2_ingest")
    workdir.mkdir(parents=True, exist_ok=True)

    log.info("plan: years=%s levels=%s", years, levels)

    overall_rc = 0

    # Multi-year runs spawn a fresh Python subprocess per year. DuckDB's
    # spatial extension has memory issues with repeated XLSX `st_read`
    # calls in a long-lived process — past ~20 transforms the process
    # SIGSEGVs. Forking per year sidesteps that without changing
    # the per-year code path.
    if len(years) > 1:
        import subprocess
        for year in years:
            log.info("=" * 70)
            log.info("=== SUBPROCESS FOR YEAR %d ===", year)
            log.info("=" * 70)
            cmd = [sys.executable, sys.argv[0], "--years", str(year)]
            if args.levels:
                cmd += ["--levels", args.levels]
            if args.workdir:
                cmd += ["--workdir", args.workdir]
            if args.r2_prefix_override:
                cmd += ["--r2-prefix-override", args.r2_prefix_override]
            if args.max_rows is not None:
                cmd += ["--max-rows", str(args.max_rows)]
            if args.skip_if_unchanged:
                cmd += ["--skip-if-unchanged"]
            if args.dry_run:
                cmd += ["--dry-run"]
            try:
                rc = subprocess.run(cmd, check=False).returncode
            except Exception:
                log.exception("[%d] subprocess orchestration failed", year)
                rc = 1
            if rc != 0:
                overall_rc = rc
                log.error("[%d] subprocess exited %d", year, rc)
        try:
            shutil.rmtree(workdir, ignore_errors=True)
        except Exception:
            pass
        return overall_rc

    # Single-year path runs in-process (DuckDB happy with one transform pass).
    with httpx.Client(headers={"User-Agent": USER_AGENT}) as client:
        for year in years:
            try:
                rc = run_one_year(
                    bls_oews_year=year,
                    client=client, workdir=workdir,
                    levels=levels,
                    max_rows=args.max_rows,
                    r2_prefix_override=args.r2_prefix_override,
                    skip_if_unchanged=args.skip_if_unchanged,
                    dry_run=args.dry_run,
                )
                if rc != 0:
                    overall_rc = rc
            except Exception:
                log.exception("[%d] year-level orchestration failed", year)
                overall_rc = 1

    try:
        shutil.rmtree(workdir, ignore_errors=True)
    except Exception:
        pass

    return overall_rc


if __name__ == "__main__":
    sys.exit(main())
