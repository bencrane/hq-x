#!/usr/bin/env python3
"""Ingest NAICS and PSC lookup datasets into lookup schema tables.

Usage:
SUPER_ADMIN_JWT_SECRET=unused-for-ingest doppler run --project data-engine-x-api --config prd -- python3 scripts/ingest_naics_psc_lookups.py
"""

from __future__ import annotations

import json
import os
import re
import sys
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

# Allow running as `python3 scripts/...` from repo root.
REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from app.config import get_settings

try:
    from psycopg import connect
except ModuleNotFoundError:
    venv_python = REPO_ROOT / ".venv/bin/python3"
    if venv_python.exists() and Path(sys.executable) != venv_python:
        os.execv(str(venv_python), [str(venv_python), *sys.argv])
    raise

NAICS_SOURCE_PATH = REPO_ROOT / "docs/api-reference-docs/naics/2022_NAICS_Structure.xlsx"
PSC_SOURCE_PATH = REPO_ROOT / "docs/api-reference-docs/psc/PSC_April_2025.xlsx"
PSC_SHEET_NAME = "PSC for 042025"
NAICS_EXPECTED_ROW_COUNT = 2125


@dataclass
class IngestSummary:
    naics_rows_upserted: int
    psc_rows_upserted: int
    naics_table_count: int
    psc_table_count: int
    naics_level_distribution: dict[int, int]


def _normalize_naics_code(raw_code: Any) -> str | None:
    if raw_code is None:
        return None
    if isinstance(raw_code, (int, float)):
        return str(int(raw_code))
    text = str(raw_code).strip()
    if not text:
        return None
    if text.endswith(".0"):
        return text[:-2]
    return text


def _clean_naics_title(raw_title: Any) -> str | None:
    if raw_title is None:
        return None
    text = str(raw_title).strip()
    if not text:
        return None
    if text.endswith("T"):
        text = text[:-1].rstrip()
    return text


def _derive_naics_sector(code: str) -> str:
    first_two = code[:2]
    if first_two in {"31", "32", "33"}:
        return "31-33"
    if first_two in {"44", "45"}:
        return "44-45"
    if first_two in {"48", "49"}:
        return "48-49"
    return first_two


def load_naics_rows() -> tuple[list[tuple[Any, ...]], dict[int, int]]:
    if not NAICS_SOURCE_PATH.exists():
        raise RuntimeError(f"NAICS source file not found: {NAICS_SOURCE_PATH}")

    workbook = load_workbook(NAICS_SOURCE_PATH, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    rows: list[tuple[Any, ...]] = []
    level_distribution: dict[int, int] = {2: 0, 3: 0, 4: 0, 5: 0, 6: 0}

    for row in worksheet.iter_rows(min_row=4, values_only=True):
        change_indicator, code_raw, title_raw = row[0], row[1], row[2]
        code = _normalize_naics_code(code_raw)
        if code is None:
            continue

        level = len(code)
        if level < 2 or level > 6:
            continue

        title = _clean_naics_title(title_raw)
        if title is None:
            continue

        sector_code = _derive_naics_sector(code)
        subsector_code = code[:3] if level >= 3 else None
        industry_group_code = code[:4] if level >= 4 else None
        industry_code = code[:5] if level >= 5 else None

        rows.append(
            (
                code,
                title,
                level,
                sector_code,
                subsector_code,
                industry_group_code,
                industry_code,
                str(change_indicator).strip() if change_indicator is not None else None,
            )
        )
        level_distribution[level] = level_distribution.get(level, 0) + 1

    if len(rows) != NAICS_EXPECTED_ROW_COUNT:
        raise RuntimeError(
            f"Expected {NAICS_EXPECTED_ROW_COUNT} NAICS rows, got {len(rows)}."
        )

    return rows, level_distribution


def _normalize_psc_code(raw_code: Any) -> str | None:
    if raw_code is None:
        return None

    if isinstance(raw_code, (int, float)):
        code = str(int(raw_code))
        return code.zfill(4)

    text = str(raw_code).strip().upper()
    if not text:
        return None

    if re.fullmatch(r"\d+(\.0+)?", text):
        code = str(int(float(text)))
        return code.zfill(4)

    return text


def _normalize_optional_date(raw_value: Any) -> date | None:
    if raw_value in (None, ""):
        return None
    if isinstance(raw_value, datetime):
        return raw_value.date()
    if isinstance(raw_value, date):
        return raw_value
    return None


def _parse_parent_code(raw_parent: Any) -> str | None:
    if raw_parent in (None, ""):
        return None
    parent_text = str(raw_parent).strip()
    if not parent_text:
        return None
    parent_prefix = parent_text.split(" - ", 1)[0].strip()
    return _normalize_psc_code(parent_prefix)


def _to_int_or_none(raw_value: Any) -> int | None:
    if raw_value in (None, ""):
        return None
    try:
        return int(float(raw_value))
    except (TypeError, ValueError):
        return None


def load_psc_rows() -> list[tuple[Any, ...]]:
    if not PSC_SOURCE_PATH.exists():
        raise RuntimeError(f"PSC source file not found: {PSC_SOURCE_PATH}")

    workbook = load_workbook(PSC_SOURCE_PATH, read_only=True, data_only=True)
    sheet_name = PSC_SHEET_NAME if PSC_SHEET_NAME in workbook.sheetnames else workbook.sheetnames[0]
    worksheet = workbook[sheet_name]

    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = {str(value).strip(): idx for idx, value in enumerate(header_row) if value is not None}

    required_headers = [
        "PSC CODE",
        "PRODUCT AND SERVICE CODE NAME",
        "START DATE",
        "END DATE",
        "PRODUCT AND SERVICE CODE FULL NAME (DESCRIPTION)",
        "PRODUCT AND SERVICE CODE INCLUDES",
        "PRODUCT AND SERVICE CODE EXCLUDES",
        "PRODUCT AND SERVICE CODE NOTES",
        "Parent PSC Code",
        "PSC Category: Service (S)/Product (P)",
        "Level 1 Category Code",
        "Level 1 Category",
        "Level 2 Category Code",
        "Level 2 Category",
    ]
    missing = [header for header in required_headers if header not in headers]
    if missing:
        raise RuntimeError(f"PSC spreadsheet is missing expected headers: {missing}")

    rows: list[tuple[Any, ...]] = []
    seen_codes: set[str] = set()
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        raw_code = row[headers["PSC CODE"]]
        psc_code = _normalize_psc_code(raw_code)
        if psc_code is None:
            continue

        end_date = _normalize_optional_date(row[headers["END DATE"]])
        if end_date is not None:
            # Current-only ingest per directive.
            continue

        if psc_code in seen_codes:
            raise RuntimeError(f"Duplicate current PSC code encountered: {psc_code}")
        seen_codes.add(psc_code)

        start_date = _normalize_optional_date(row[headers["START DATE"]])
        parent_psc_code = _parse_parent_code(row[headers["Parent PSC Code"]])
        psc_name = row[headers["PRODUCT AND SERVICE CODE NAME"]]
        psc_full_name = row[headers["PRODUCT AND SERVICE CODE FULL NAME (DESCRIPTION)"]]

        raw_category = row[headers["PSC Category: Service (S)/Product (P)"]]
        if raw_category is not None and str(raw_category).strip().upper().startswith("P"):
            is_product = True
            psc_category = "Product"
        elif raw_category is not None and str(raw_category).strip().upper().startswith("S"):
            is_product = False
            psc_category = "Service"
        else:
            is_product = psc_code[:1].isdigit()
            psc_category = "Product" if is_product else "Service"

        rows.append(
            (
                psc_code,
                str(psc_name).strip() if psc_name is not None else None,
                str(psc_full_name).strip() if psc_full_name is not None else None,
                parent_psc_code,
                psc_category,
                is_product,
                _to_int_or_none(row[headers["Level 1 Category Code"]]),
                str(row[headers["Level 1 Category"]]).strip()
                if row[headers["Level 1 Category"]] is not None
                else None,
                _to_int_or_none(row[headers["Level 2 Category Code"]]),
                str(row[headers["Level 2 Category"]]).strip()
                if row[headers["Level 2 Category"]] is not None
                else None,
                start_date,
                end_date,
                True,
                str(row[headers["PRODUCT AND SERVICE CODE INCLUDES"]]).strip()
                if row[headers["PRODUCT AND SERVICE CODE INCLUDES"]] is not None
                else None,
                str(row[headers["PRODUCT AND SERVICE CODE EXCLUDES"]]).strip()
                if row[headers["PRODUCT AND SERVICE CODE EXCLUDES"]] is not None
                else None,
                str(row[headers["PRODUCT AND SERVICE CODE NOTES"]]).strip()
                if row[headers["PRODUCT AND SERVICE CODE NOTES"]] is not None
                else None,
            )
        )

    return rows


def upsert_lookup_rows(
    naics_rows: list[tuple[Any, ...]],
    psc_rows: list[tuple[Any, ...]],
    naics_level_distribution: dict[int, int],
) -> IngestSummary:
    settings = get_settings()

    naics_sql = """
    INSERT INTO lookup.naics_codes (
        naics_code,
        title,
        level,
        sector_code,
        subsector_code,
        industry_group_code,
        industry_code,
        change_indicator,
        updated_at
    )
    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, NOW())
    ON CONFLICT (naics_code)
    DO UPDATE SET
        title = EXCLUDED.title,
        level = EXCLUDED.level,
        sector_code = EXCLUDED.sector_code,
        subsector_code = EXCLUDED.subsector_code,
        industry_group_code = EXCLUDED.industry_group_code,
        industry_code = EXCLUDED.industry_code,
        change_indicator = EXCLUDED.change_indicator,
        updated_at = NOW()
    """

    psc_sql = """
    INSERT INTO lookup.psc_codes (
        psc_code,
        psc_name,
        psc_full_name,
        parent_psc_code,
        psc_category,
        is_product,
        level_1_category_code,
        level_1_category,
        level_2_category_code,
        level_2_category,
        start_date,
        end_date,
        is_current,
        includes_text,
        excludes_text,
        notes_text,
        updated_at
    )
    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
    ON CONFLICT (psc_code)
    DO UPDATE SET
        psc_name = EXCLUDED.psc_name,
        psc_full_name = EXCLUDED.psc_full_name,
        parent_psc_code = EXCLUDED.parent_psc_code,
        psc_category = EXCLUDED.psc_category,
        is_product = EXCLUDED.is_product,
        level_1_category_code = EXCLUDED.level_1_category_code,
        level_1_category = EXCLUDED.level_1_category,
        level_2_category_code = EXCLUDED.level_2_category_code,
        level_2_category = EXCLUDED.level_2_category,
        start_date = EXCLUDED.start_date,
        end_date = EXCLUDED.end_date,
        is_current = EXCLUDED.is_current,
        includes_text = EXCLUDED.includes_text,
        excludes_text = EXCLUDED.excludes_text,
        notes_text = EXCLUDED.notes_text,
        updated_at = NOW()
    """

    with connect(settings.database_url) as conn:
        with conn.cursor() as cur:
            cur.executemany(naics_sql, naics_rows)
            cur.executemany(psc_sql, psc_rows)
            conn.commit()

            cur.execute("SELECT COUNT(*) FROM lookup.naics_codes")
            naics_count = int(cur.fetchone()[0])

            cur.execute("SELECT COUNT(*) FROM lookup.psc_codes")
            psc_count = int(cur.fetchone()[0])

    return IngestSummary(
        naics_rows_upserted=len(naics_rows),
        psc_rows_upserted=len(psc_rows),
        naics_table_count=naics_count,
        psc_table_count=psc_count,
        naics_level_distribution=naics_level_distribution,
    )


def main() -> None:
    naics_rows, naics_level_distribution = load_naics_rows()
    psc_rows = load_psc_rows()
    summary = upsert_lookup_rows(naics_rows, psc_rows, naics_level_distribution)

    print(
        json.dumps(
            {
                "naics_rows_upserted": summary.naics_rows_upserted,
                "psc_rows_upserted": summary.psc_rows_upserted,
                "naics_table_count": summary.naics_table_count,
                "psc_table_count": summary.psc_table_count,
                "naics_level_distribution": summary.naics_level_distribution,
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
